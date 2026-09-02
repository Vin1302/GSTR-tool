from __future__ import annotations

import json
import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from gstr_tool.core.credentials import load_credentials
from gstr_tool.core.browser import GstBrowserSession, financial_year_periods
from gstr_tool.core.parsers import parse_download_folder
from gstr_tool.core.pdf_reader import parse_gstr3b_text
from gstr_tool.core.template_writer import generate_workbook


# The template holds real client data, so it is not committed. Point
# GSTR_TEMPLATE at a copy to run the template tests.
TEMPLATE = Path(os.environ.get("GSTR_TEMPLATE", "")) if os.environ.get("GSTR_TEMPLATE") else None


PORTAL_2B_HEADER = [
    "GSTIN of supplier", "Trade/Legal name", "Invoice number", "Invoice type", "Invoice Date",
    "Invoice Value(₹)", "Place of supply", "Supply Attract Reverse Charge", "Rate(%)",
    "Taxable Value (₹)", "Integrated Tax(₹)", "Central Tax(₹)", "State/UT Tax(₹)", "Cess(₹)",
    "GSTR-1/IFF/GSTR-5 Period", "ITC Availability", "Reason",
]
PORTAL_CDNR_HEADER = [
    "GSTIN of supplier", "Trade/Legal name", "Note number", "Note type", "Note Supply type",
    "Note date", "Note Value(₹)", "Place of supply", "Supply Attract Reverse Charge", "Rate(%)",
    "Taxable Value (₹)", "Integrated Tax(₹)", "Central Tax(₹)", "State/UT Tax(₹)", "Cess(₹)",
    "GSTR-1/IFF/GSTR-5 Period", "ITC Availability", "Reason",
]
PORTAL_EINVOICE_HEADER = [
    "Irn", "Irn date", "Document type", "Document number", "Document date",
    "GSTIN of recipient", "Recipient name", "Document value", "Taxable value",
    "Integrated tax", "Central tax", "State/UT tax", "Cess", "Status",
]


def _portal_2b_workbook(path: Path) -> None:
    """A GSTR-2B export shaped like GSTN's: title rows, B2B and B2B-CDNR sheets."""
    workbook = Workbook()
    b2b = workbook.active
    b2b.title = "B2B"
    b2b.append(["Goods and Services Tax - GSTR-2B"])
    b2b.append(["GSTIN of Recipient", "19ZZZZZ0000Z1Z5"])
    b2b.append(["Period", "Apr-25"])
    b2b.append([])
    b2b.append(PORTAL_2B_HEADER)
    b2b.append(["19AAAAA0000A1Z5", "Vendor A", "P-1", "R", "07/04/2025", 590, "19-WB", "N", 18,
                500, 90, 0, 0, 0, "042025", "Yes", ""])
    b2b.append(["19BBBBB0000B1Z5", "Vendor B", "P-2", "R", "08/04/2025", 826, "19-WB", "Y", 18,
                700, 0, 63, 63, 0, "042025", "Yes", ""])
    b2b.append(["19CCCCC0000C1Z5", "Vendor C", "P-3", "R", "09/04/2025", 1062, "19-WB", "N", 18,
                900, 162, 0, 0, 0, "042025", "No", "POS and supplier state are same"])

    cdnr = workbook.create_sheet("B2B-CDNR")
    cdnr.append(["Goods and Services Tax - GSTR-2B"])
    cdnr.append(PORTAL_CDNR_HEADER)
    cdnr.append(["19AAAAA0000A1Z5", "Vendor A", "CN-1", "C", "R", "10/04/2025", 118, "19-WB",
                 "N", 18, 100, 18, 0, 0, 0, "042025", "Yes", ""])
    workbook.save(path)


def _portal_einvoice_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "e-Invoice"
    sheet.append(["Details from e-Invoices"])
    sheet.append(PORTAL_EINVOICE_HEADER)
    sheet.append(["IRN1", "05/04/2025", "INV", "S-1", "05/04/2025", "19AAAAA0000A1Z5",
                  "Buyer A", 1180, 1000, 180, 0, 0, 0, "Active"])
    sheet.append(["IRN2", "06/04/2025", "INV", "S-2", "06/04/2025", "19AAAAA0000A1Z5",
                  "Buyer A", 9999, 8474, 1525, 0, 0, 0, "Cancelled"])
    sheet.append(["IRN3", "07/04/2025", "CRN", "C-1", "07/04/2025", "19AAAAA0000A1Z5",
                  "Buyer A", 236, 200, 36, 0, 0, 0, "Active"])
    workbook.save(path)


GSTR3B_TEXT = """Form GSTR-3B
Year 2025-26 Period April
3.1 Details of Outward supplies and inward supplies liable to reverse charge
(a) Outward taxable supplies (other than zero rated, nil rated and exempted) 2000.00 360.00 0.00 0.00 0.00
(b) Outward taxable supplies (zero rated) 0.00 0.00 0.00 0.00 0.00
(c) Other outward supplies (Nil rated, exempted) 100.00 0.00 0.00 0.00 0.00
(d) Inward supplies (liable to reverse charge) 500.00 90.00 0.00 0.00 0.00
(e) Non-GST outward supplies 25.00 0.00 0.00 0.00 0.00
4. Eligible ITC
(1) Import of goods 10.00 0.00 0.00 0.00
(2) Import of services 0.00 0.00 0.00 0.00
(3) Inward supplies liable to reverse charge (other than 1 & 2 above) 90.00 0.00 0.00 0.00
(5) All other ITC 700.00 20.00 20.00 0.00
(B) ITC Reversed
(1) As per rules 42 & 43 of CGST Rules 0.00 5.00 5.00 0.00
(2) Others 0.00 1.00 1.00 0.00
5.1 Interest and Late fee
Late fee 0.00 50.00 50.00 0.00
"""


class GstrToolTests(unittest.TestCase):
    def test_financial_year_periods(self):
        periods = financial_year_periods("2025-26")
        self.assertEqual(periods[0], ("April", "1", "Apr-2025"))
        self.assertEqual(periods[-1], ("March", "4", "Mar-2026"))
        self.assertEqual(len(periods), 12)

    def test_credentials_aliases(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "credentials.xlsx"
            workbook = Workbook(); sheet = workbook.active
            sheet.append(["Client Name", "GSTIN", "User ID", "Password"])
            sheet.append(["Acme", "19ABCDE1234F1Z5", "acme.user", "secret"])
            workbook.save(path)
            credentials = load_credentials(path)
            self.assertEqual(credentials[0].label, "Acme")
            self.assertEqual(credentials[0].username, "acme.user")

    def test_client_root_never_nests_on_repeated_runs(self):
        with tempfile.TemporaryDirectory() as folder:
            # resolve() first: Windows hands out 8.3 short temp paths
            # (RUNNER~1), and resolve_client_root returns the long form.
            base = Path(folder).resolve()
            first = GstBrowserSession.resolve_client_root(base, "Acme Pvt Ltd", "2025-26")
            self.assertEqual(first, base / "Acme_Pvt_Ltd" / "2025-26")
            # Re-running from the folder the previous run produced, or from one
            # of its report folders, must land on the same place.
            self.assertEqual(GstBrowserSession.resolve_client_root(first, "Acme Pvt Ltd", "2025-26"), first)
            self.assertEqual(
                GstBrowserSession.resolve_client_root(first / "GSTR-2B", "Acme Pvt Ltd", "2025-26"), first)
            self.assertEqual(
                GstBrowserSession.resolve_client_root(first.parent, "Acme Pvt Ltd", "2025-26"), first)
            # A different financial year stays a sibling, not a child.
            self.assertEqual(GstBrowserSession.resolve_client_root(first, "Acme Pvt Ltd", "2024-25"),
                             base / "Acme_Pvt_Ltd" / "2024-25")

    def test_existing_downloads_are_not_repeated(self):
        with tempfile.TemporaryDirectory() as folder:
            session = GstBrowserSession(folder)
            report_folder = Path(folder)
            self.assertIsNone(session._existing_download(report_folder, "GSTR-2B", "Apr-2025", "excel"))
            (report_folder / "Apr-2025_GSTR2B_excel_portal.xlsx").write_text("x")
            found = session._existing_download(report_folder, "GSTR-2B", "Apr-2025", "excel")
            self.assertIsNotNone(found)
            self.assertIsNone(session._existing_download(report_folder, "GSTR-2B", "May-2025", "excel"))

    def test_stray_downloads_are_filed_by_name(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            session = GstBrowserSession(root)
            session.staging = root / "staging"
            session.staging.mkdir()
            session._report_folders = {name: root / name for name in
                                       ("GSTR-1", "GSTR-3B", "GSTR-2B", "E-Invoice")}
            for target in session._report_folders.values():
                target.mkdir()
            # A file GSTN delivered after its step had moved on.
            (session.staging / "einvoice_details_042025.xlsx").write_text("x")
            (session.staging / "GSTR2B_042025.xlsx").write_text("x")
            (session.staging / "unknown_thing.txt").write_text("x")

            filed = session.sweep_staging()

            self.assertEqual(len(filed), 2)
            self.assertTrue((root / "E-Invoice" / "einvoice_details_042025.xlsx").exists())
            self.assertTrue((root / "GSTR-2B" / "GSTR2B_042025.xlsx").exists())
            # Anything unrecognised stays put rather than being filed wrongly.
            self.assertTrue((session.staging / "unknown_thing.txt").exists())

    def test_download_directory_is_reasserted_before_every_click(self):
        """A page navigation can drop Chrome's download directory.

        Setting it once per run silently sent every file to Chrome's own
        default folder, where nothing was watching for it.
        """
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            session = GstBrowserSession(root)
            session.staging = root / "staging"
            session.staging.mkdir()
            session._report_folders = {}
            sent = []

            class FakeDriver:
                def execute_cdp_cmd(self, command, params):
                    sent.append((command, params["downloadPath"]))

            session.driver = FakeDriver()
            session._prepare_click()
            session._prepare_click()

            commands = [command for command, _ in sent]
            self.assertIn("Browser.setDownloadBehavior", commands)
            # Once per click, not once per run.
            self.assertEqual(commands.count("Browser.setDownloadBehavior"), 2)
            self.assertTrue(all(path == str(session.staging.resolve()) for _, path in sent))

    def test_scoped_action_stops_at_a_block_naming_another_report(self):
        """The tile is the block naming this report and no other.

        Three tiles carry a button reading VIEW, so a container that also
        mentions GSTR-1 is too big to be the GSTR-2B tile — which is how every
        report ended up clicking GSTR-1's button.
        """
        class FakeNode:
            def __init__(self, text, parent=None, clickable=False):
                self.text = text
                self.parent = parent
                self.clickable = clickable

            def find_elements(self, by, xpath):
                return [self.parent] if xpath == ".." and self.parent else []

        page = FakeNode("Details of outward supplies GSTR-1 VIEW DOWNLOAD "
                        "Auto - drafted ITC Statement GSTR-2B VIEW DOWNLOAD", clickable=True)
        tile = FakeNode("Auto - drafted ITC Statement for the month GSTR-2B VIEW DOWNLOAD",
                        parent=page, clickable=True)
        header = FakeNode("Auto - drafted ITC Statement for the month GSTR-2B", parent=tile)
        heading = FakeNode("GSTR-2B", parent=header)

        session = GstBrowserSession("/tmp")
        session._heading_element = lambda labels: heading
        clicked = []
        session._click_text = lambda labels, root=None: (
            clicked.append(root) or True) if getattr(root, "clickable", False) else False

        opened = session._click_scoped_action(
            ["gstr-2b"], ["view"], ["gstr-1", "gstr-3b", "gstr-2a"])

        self.assertTrue(opened)
        self.assertIs(clicked[0], tile, "clicked outside the GSTR-2B tile")

    def test_scoped_action_refuses_when_only_the_whole_page_matches(self):
        class FakeNode:
            def __init__(self, text, parent=None):
                self.text = text
                self.parent = parent

            def find_elements(self, by, xpath):
                return [self.parent] if xpath == ".." and self.parent else []

        page = FakeNode("GSTR-1 VIEW GSTR-2B VIEW")
        heading = FakeNode("GSTR-2B", parent=page)
        session = GstBrowserSession("/tmp")
        session._heading_element = lambda labels: heading
        session._click_text = lambda labels, root=None: True

        self.assertFalse(session._click_scoped_action(
            ["gstr-2b"], ["view"], ["gstr-1", "gstr-3b", "gstr-2a"]))

    def test_summary_separates_skipped_from_failed(self):
        results = [
            "GSTR-1 PDF Apr-2025: Apr-2025_GSTR1_pdf_x.pdf",
            "GSTR-1 PDF May-2025: already downloaded (May-2025_GSTR1_pdf_x.pdf)",
            "GSTR-2B EXCEL Apr-2025: tile not available for this period",
        ]
        summary = " ".join(GstBrowserSession._summarise(results))
        self.assertIn("1 file(s) downloaded", summary)
        self.assertIn("1 already present and skipped", summary)
        self.assertIn("1 not obtained", summary)

    def test_tile_action_picks_the_button_under_its_own_heading(self):
        """Laid out as the portal lays the dashboard out, from the recording.

        Every report was opening GSTR-1 because the tile lookup returned a
        container holding all the tiles, whose first button is GSTR-1's. The
        button is now chosen by position, so GSTR-2B's heading reaches GSTR-2B's
        VIEW even though three tiles carry a button with that same label.
        """
        class FakeButton:
            def __init__(self, text, x, y):
                self.text = text
                self.rect = {"x": x, "y": y, "width": 90, "height": 30}

            def is_displayed(self):
                return True

            def is_enabled(self):
                return True

            def get_attribute(self, name):
                return None

        class FakeHeading:
            def __init__(self, x, y):
                self.rect = {"x": x, "y": y, "width": 200, "height": 40}

            def is_displayed(self):
                return True

        # Row 1: GSTR-1 at x=210, GSTR-2B at x=660. Row 2: GSTR-3B at x=210.
        buttons = [
            FakeButton("VIEW", 240, 145), FakeButton("DOWNLOAD", 350, 145),
            FakeButton("VIEW", 690, 145), FakeButton("DOWNLOAD", 770, 145),
            FakeButton("VIEW GSTR3B", 220, 300), FakeButton("DOWNLOAD", 360, 300),
        ]
        session = GstBrowserSession("/tmp")
        session._heading_element = lambda labels: FakeHeading(660, 70)
        session._action_candidates = lambda labels: [
            (button, " ".join(button.text.split()).lower() in labels)
            for button in buttons
            if any(label in button.text.lower() for label in labels)
        ]
        clicked = []
        session._click_element = clicked.append

        self.assertTrue(session._tile_action(["gstr-2b"], ["view", "download"]))
        # GSTR-2B's own VIEW, not GSTR-1's identical one further left.
        self.assertEqual(clicked[0].rect["x"], 690)
        self.assertEqual(clicked[0].text, "VIEW")

    def test_tile_action_ignores_buttons_above_the_heading(self):
        """A tile's button never sits above its heading, but another tile's does."""
        class FakeButton:
            def __init__(self, text, x, y):
                self.text = text
                self.rect = {"x": x, "y": y, "width": 90, "height": 30}

        class FakeHeading:
            rect = {"x": 210, "y": 260, "width": 200, "height": 40}

        session = GstBrowserSession("/tmp")
        session._heading_element = lambda labels: FakeHeading()
        # The only candidate belongs to the tile in the row above.
        session._action_candidates = lambda labels: [(FakeButton("VIEW", 240, 145), True)]
        session._click_element = lambda element: self.fail("clicked another tile's button")

        self.assertFalse(session._tile_action(["gstr-3b"], ["view"]))

    def test_tile_lookup_walks_up_from_the_deepest_match(self):
        """The tile is the block around the heading, never a page-wide wrapper.

        Matching every element that contains the text also matches the outer
        wrappers, which come first in document order. Taking that match handed
        back a container holding every tile, so clicking "DOWNLOAD" inside it
        hit GSTR-1's button and every report drove the GSTR-1 tile.
        """
        class FakeElement:
            def __init__(self, name, parent=None, buttons=()):
                self.name = name
                self.parent = parent
                self.buttons = list(buttons)

            def is_displayed(self):
                return True

            def is_enabled(self):
                return True

            def find_elements(self, by, xpath):
                if xpath == "..":
                    return [self.parent] if self.parent else []
                return self.buttons

        page = FakeElement("page-wrapper", buttons=[FakeElement("GSTR-1 DOWNLOAD")])
        tile = FakeElement("gstr-2b-tile", parent=page,
                           buttons=[FakeElement("GSTR-2B DOWNLOAD")])
        heading = FakeElement("heading", parent=tile)

        session = GstBrowserSession("/tmp")
        found = session._tile_container(heading)

        self.assertIs(found, tile)
        self.assertEqual([button.name for button in found.buttons], ["GSTR-2B DOWNLOAD"])

    def test_report_is_guessed_from_the_portal_file_name(self):
        guess = GstBrowserSession._guess_report
        self.assertEqual(guess("einvoice_042025.xlsx"), "E-Invoice")
        self.assertEqual(guess("GSTR-2B_19ABCDE1234F1Z5_042025.xlsx"), "GSTR-2B")
        self.assertEqual(guess("GSTR3B_19ABCDE1234F1Z5_042025.pdf"), "GSTR-3B")
        self.assertEqual(guess("GSTR1_Summary_042025.pdf"), "GSTR-1")
        self.assertEqual(guess("statement.csv"), "")

    def test_gstr3b_pdf_text_buckets(self):
        buckets = parse_gstr3b_text(GSTR3B_TEXT)
        self.assertEqual(float(buckets["outward_nrc"].taxable), 2000)
        self.assertEqual(float(buckets["outward_nrc"].igst), 360)
        # Nil rated/exempt and non-GST supplies share the Non-Taxable column.
        self.assertEqual(float(buckets["non_taxable"].taxable), 125)
        self.assertEqual(float(buckets["outward_rcm"].taxable), 500)
        self.assertEqual(float(buckets["itc_other"].igst), 700)
        # Import of goods/services plus reverse-charge inward supplies.
        self.assertEqual(float(buckets["itc_rcm"].igst), 100)
        self.assertEqual(float(buckets["itc_reversed"].cgst), 6)
        self.assertEqual(float(buckets["late_fee"].cgst), 50)

    def test_portal_downloads_are_parsed_from_report_folders(self):
        with tempfile.TemporaryDirectory() as folder:
            downloads = Path(folder)
            for name in ("GSTR-1", "GSTR-3B", "GSTR-2B", "E-Invoice"):
                (downloads / name).mkdir()
            _portal_einvoice_workbook(downloads / "E-Invoice" / "Apr-2025_EInvoice_portal.xlsx")
            _portal_2b_workbook(downloads / "GSTR-2B" / "Apr-2025_GSTR2B_excel_portal.xlsx")
            (downloads / "GSTR-1" / "Apr-2025_GSTR1_json_returns.json").write_text(json.dumps({
                "fp": "042025",
                "b2b": [{"ctin": "19AAAAA0000A1Z5", "inv": [
                    {"inum": "S-1", "idt": "10-04-2025", "itms": [{"itm_det": {"txval": 1500, "iamt": 270}}]}]}],
            }))

            months, invoices, used = parse_download_folder(downloads)

            # e-Invoice: cancelled row dropped, credit note in its own bucket.
            self.assertEqual(float(months["April"].einvoice["b2b"].taxable), 1000)
            self.assertEqual(float(months["April"].einvoice["credit_note"].taxable), 200)
            # GSTR-2B: portal column names recognised, ineligible row dropped,
            # reverse-charge split kept, credit note negative.
            self.assertEqual(float(months["April"].gstr2b["other"].taxable), 400)
            self.assertEqual(float(months["April"].gstr2b["rcm"].taxable), 700)
            self.assertEqual(len(invoices), 3)
            credit = [item for item in invoices if item.document_type == "CREDIT NOTE"]
            self.assertEqual(len(credit), 1)
            self.assertEqual(float(credit[0].taxable), -100)
            self.assertEqual(float(credit[0].document_value), -118)
            self.assertEqual(float(months["April"].gstr1["b2b"].taxable), 1500)
            self.assertEqual(len(used), 3)

    def test_gstr3b_pdf_fills_the_3b_sheet(self):
        if TEMPLATE is None or not TEMPLATE.exists():
            self.skipTest("Set GSTR_TEMPLATE to the reconciliation template to run this test")
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            downloads = root / "downloads"
            (downloads / "GSTR-3B").mkdir(parents=True)
            _write_pdf(downloads / "GSTR-3B" / "Apr-2025_GSTR3B_pdf_filed.pdf", GSTR3B_TEXT)
            months, invoices, used = parse_download_folder(downloads)
            self.assertTrue(used, "the GSTR-3B PDF was not read")
            output = root / "result.xlsx"
            generate_workbook(TEMPLATE, output, months, invoices)
            result = load_workbook(output, data_only=False)
            sheet = result["As per 3B"]
            self.assertEqual(sheet["B7"].value, 2000)     # outward (NRC) taxable
            self.assertEqual(sheet["C7"].value, 125)      # non-taxable
            self.assertEqual(sheet["F7"].value, 360)      # IGST liability
            self.assertEqual(sheet["L7"].value, 500)      # outward (RC) taxable
            self.assertEqual(sheet["B24"].value, 700)     # ITC eligible IGST
            self.assertEqual(sheet["F24"].value, 0)       # ITC eligible CESS sits in F
            self.assertEqual(sheet["H24"].value, 6)       # ITC reversed CGST sits in H
            self.assertEqual(sheet["B60"].value, 100)     # ITC-RCM IGST, April is row 60
            self.assertEqual(sheet["S60"].value, 50)      # late fee CGST
            result.close()

    def test_filters_and_formula_preservation(self):
        if TEMPLATE is None or not TEMPLATE.exists():
            self.skipTest("Set GSTR_TEMPLATE to the reconciliation template to run this test")
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder); downloads = root / "downloads"
            (downloads / "E-Invoice").mkdir(parents=True)
            (downloads / "GSTR-2B").mkdir(parents=True)
            _portal_einvoice_workbook(downloads / "E-Invoice" / "Apr-2025_EInvoice_portal.xlsx")
            _portal_2b_workbook(downloads / "GSTR-2B" / "Apr-2025_GSTR2B_excel_portal.xlsx")

            months, invoices, used = parse_download_folder(downloads)
            output = root / "result.xlsx"
            generate_workbook(TEMPLATE, output, months, invoices)
            source = load_workbook(TEMPLATE, data_only=False)
            result = load_workbook(output, data_only=False)
            self.assertEqual(result["As per E-Invoice"]["B6"].value, 1000)
            self.assertEqual(result["GSTR 2b"]["C6"].value, 72)   # 90 IGST less the 18 credit note
            self.assertEqual(result["GSTR 2b"]["G6"].value, 700)
            # Formulas on the untouched sheets survive.
            self.assertEqual(result["As per E-Invoice"]["N6"].value, source["As per E-Invoice"]["N6"].value)
            self.assertEqual(result["Summary"]["B7"].value, source["Summary"]["B7"].value)
            source.close(); result.close()


def _write_pdf(path: Path, text: str) -> None:
    """Render plain text as a one-page PDF for the PDF-reader tests."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:  # pragma: no cover - reportlab is a test-only helper
        raise unittest.SkipTest("reportlab is needed to build the test PDF")
    page = canvas.Canvas(str(path), pagesize=A4)
    page.setFont("Helvetica", 7)
    y = 800
    for line in text.splitlines():
        page.drawString(20, y, line)
        y -= 12
    page.save()


if __name__ == "__main__":
    unittest.main()
