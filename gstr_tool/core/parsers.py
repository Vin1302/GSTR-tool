from __future__ import annotations

import json
import re
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .helpers import decimal, month_name, normalized_header, yes
from .models import Invoice2B, MonthData, TaxAmounts
from .pdf_reader import parse_gstr1_summary_pdf, parse_gstr3b_pdf


# The GST Portal downloads and hand-made workbooks spell the same column in
# several ways. Keep every spelling in one place so each parser stays readable.
TAXABLE_HEADERS = ("taxable value", "taxable amount", "taxable value rs", "total taxable value")
IGST_HEADERS = ("integrated tax", "igst", "igst amount", "integrated tax rs", "integrated tax amount")
CGST_HEADERS = ("central tax", "cgst", "cgst amount", "central tax rs", "central tax amount")
SGST_HEADERS = ("state ut tax", "state tax", "sgst", "sgst amount", "state ut tax rs")
CESS_HEADERS = ("cess", "cess amount", "cess rs")
DOCUMENT_NUMBER_HEADERS = ("invoice number", "document number", "note number", "inum",
                           "invoice no", "note no", "voucher number")
DOCUMENT_DATE_HEADERS = ("invoice date", "document date", "note date", "date", "note supply date")
DOCUMENT_TYPE_HEADERS = ("document type", "note type", "invoice type", "type")
DOCUMENT_VALUE_HEADERS = ("invoice value", "document value", "note value", "total invoice value")


def _amounts(record: dict[str, Any]) -> TaxAmounts:
    return TaxAmounts(
        taxable=decimal(record.get("txval", record.get("taxable_value", record.get("taxable")))),
        igst=decimal(record.get("iamt", record.get("igst"))),
        cgst=decimal(record.get("camt", record.get("cgst"))),
        sgst=decimal(record.get("samt", record.get("sgst"))),
        cess=decimal(record.get("csamt", record.get("cess"))),
    )


def _bucket(target: dict[str, TaxAmounts], name: str) -> TaxAmounts:
    return target.setdefault(name, TaxAmounts())


def _walk_items(invoice: dict[str, Any]) -> TaxAmounts:
    result = TaxAmounts()
    for item in invoice.get("itms", invoice.get("items", [])) or []:
        details = item.get("itm_det", item)
        result.add(_amounts(details))
    if not invoice.get("itms") and not invoice.get("items"):
        result.add(_amounts(invoice))
    return result


def parse_gstr1_json(path: str | Path, months: dict[str, MonthData]) -> None:
    payload = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    root = payload.get("data", payload)
    fallback = month_name("", str(root.get("fp", "")))
    sections = {
        "b2b": "b2b", "b2ba": "b2b_amended", "b2cs": "b2c", "b2cl": "b2c",
        "b2csa": "b2c_amended", "b2cla": "b2c_amended", "cdnr": "cdn", "cdnur": "cdn",
    }
    for section, category in sections.items():
        for supplier in root.get(section, []) or []:
            invoices = supplier.get("inv", supplier.get("nt", [supplier])) or []
            for invoice in invoices:
                month = month_name(invoice.get("idt", invoice.get("nt_dt", "")), fallback)
                if not month:
                    continue
                sign = -1 if str(invoice.get("ntty", invoice.get("typ", ""))).upper().startswith("C") else 1
                if str(invoice.get("rchrg", "")).upper() == "Y" and category == "b2b":
                    category_name = "b2b_rcm"
                else:
                    category_name = category
                _bucket(months[month].gstr1, category_name).add(_walk_items(invoice), sign)
    for row in root.get("hsn", {}).get("data", root.get("hsn", [])) or []:
        _bucket(months[fallback].gstr1, "hsn").add(_amounts(row))


def parse_gstr3b_json(path: str | Path, months: dict[str, MonthData]) -> None:
    payload = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    root = payload.get("data", payload)
    month = month_name("", str(root.get("ret_period", root.get("fp", ""))))
    if not month:
        return
    target = months[month].gstr3b
    sup = root.get("sup_details", root.get("supDetails", {}))
    _bucket(target, "outward_nrc").add(_amounts(sup.get("osup_det", {})))
    _bucket(target, "outward_rcm").add(_amounts(sup.get("isup_rev", {})))
    _bucket(target, "non_taxable").add(_amounts(sup.get("osup_nil_exmp", {})))
    itc = root.get("itc_elg", {})
    for row in itc.get("itc_avl", []) or []:
        key = "itc_rcm" if str(row.get("ty", "")).upper() in {"IMPG", "ISRC", "RCM"} else "itc_other"
        _bucket(target, key).add(_amounts(row))
    for row in itc.get("itc_rev", []) or []:
        _bucket(target, "itc_reversed").add(_amounts(row))
    for row in root.get("intr_ltfee", {}).get("intr_details", []) or []:
        _bucket(target, "late_fee").add(_amounts(row))


def _is_header_row(values: list[str]) -> bool:
    """Recognise the header line of a GST Portal sheet or a hand-made workbook.

    Portal sheets carry several title rows above the header, and the credit-note
    sheets label their key column "Note number" rather than "Invoice number", so
    the header is found by content rather than by a fixed row index.
    """
    present = set(values)
    if present.intersection(DOCUMENT_NUMBER_HEADERS):
        return True
    return bool(present.intersection(TAXABLE_HEADERS)) and bool(
        present.intersection(IGST_HEADERS + CGST_HEADERS)
    )


def _rows_from_excel(path: Path) -> Iterable[tuple[str, dict[str, Any]]]:
    """Yield ``(sheet name, record)`` for every data row of every sheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = None
            for row in rows:
                values = [normalized_header(value) for value in row]
                if _is_header_row(values):
                    header = values
                    break
            if not header:
                continue
            for row in rows:
                if not any(value not in (None, "") for value in row):
                    continue
                record = {header[index]: value for index, value in enumerate(row)
                          if index < len(header) and header[index]}
                yield sheet.title, record
    finally:
        workbook.close()


def _pick(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return None


def _row_amounts(row: dict[str, Any], sign: int = 1) -> TaxAmounts:
    amounts = TaxAmounts(
        taxable=decimal(_pick(row, *TAXABLE_HEADERS)),
        igst=decimal(_pick(row, *IGST_HEADERS)),
        cgst=decimal(_pick(row, *CGST_HEADERS)),
        sgst=decimal(_pick(row, *SGST_HEADERS)),
        cess=decimal(_pick(row, *CESS_HEADERS)),
    )
    if sign == 1:
        return amounts
    negated = TaxAmounts()
    negated.add(amounts, sign)
    return negated


def _document_category(sheet_name: str, document_type: str) -> tuple[str, int]:
    """Classify a row as invoice/debit note/credit note and give its sign.

    Credit notes reduce the figures they belong to, so they are stored negative.
    GSTN uses both words ("Credit Note") and codes (CRN/DBN, C/D).
    """
    text = f"{sheet_name} {document_type}".lower()
    if "credit" in text or re.search(r"\bcrn\b|\bcdnr?a?\b.*\bc\b", text) or text.strip() in {"c", "crn"}:
        if "debit" not in text:
            return "credit_note", -1
    if "debit" in text or text.strip() in {"d", "dbn"} or re.search(r"\bdbn\b", text):
        return "debit_note", 1
    return "b2b", 1


def _is_cancelled(row: dict[str, Any]) -> bool:
    """e-Invoice rows are dropped only when GSTN reports them cancelled or invalid.

    Not every export carries a status column, so a missing status means the row
    counts; only an explicit cancellation removes it.
    """
    status = str(_pick(row, "status", "invoice status", "irn status", "e invoice status",
                       "einvoice status", "cancel status") or "").strip().lower()
    if not status:
        return False
    if status in {"valid", "active", "live", "generated"}:
        return False
    return any(term in status for term in ("cancel", "invalid", "inactive", "deleted"))


def parse_einvoice_excel(path: str | Path, months: dict[str, MonthData],
                         fallback_period: str = "") -> None:
    for sheet_name, row in _rows_from_excel(Path(path)):
        if _is_cancelled(row):
            continue
        date_value = _pick(row, *DOCUMENT_DATE_HEADERS)
        month = month_name(date_value) or month_name("", fallback_period)
        if not month:
            continue
        document_type = str(_pick(row, *DOCUMENT_TYPE_HEADERS) or "invoice")
        category, _sign = _document_category(sheet_name, document_type)
        # The e-invoice sheet gives credit and debit notes their own columns and
        # the template's TOTAL subtracts the credit-note one, so all three
        # categories are accumulated as positive figures.
        _bucket(months[month].einvoice, category).add(_row_amounts(row))


def parse_2b_excel(path: str | Path, months: dict[str, MonthData], invoices: list[Invoice2B],
                   fallback_period: str = "") -> None:
    for sheet_name, row in _rows_from_excel(Path(path)):
        raw_availability = _pick(row, "itc availability", "itc available", "itc availability status")
        if raw_availability is not None and not yes(raw_availability):
            continue
        raw_reverse = _pick(row, "supply attract reverse charge", "reverse charge", "rcm")
        reverse = yes(raw_reverse)
        period = _pick(row, "2b period", "return period", "tax period")
        date_value = _pick(row, *DOCUMENT_DATE_HEADERS)
        # The statement's own period wins: a 2B for July can carry a June invoice.
        month = month_name("", fallback_period) or month_name(period) or month_name(date_value)
        if not month:
            continue
        document_type = str(_pick(row, *DOCUMENT_TYPE_HEADERS) or "INVOICE")
        category, sign = _document_category(sheet_name, document_type)
        values = _row_amounts(row, sign)
        _bucket(months[month].gstr2b, "rcm" if reverse else "other").add(values)
        document_value = decimal(_pick(row, *DOCUMENT_VALUE_HEADERS))
        invoices.append(Invoice2B(
            party=str(_pick(row, "trade legal name", "trade name", "legal name", "supplier name", "party") or ""),
            gstin=str(_pick(row, "gstin of supplier", "supplier gstin", "gstin") or ""),
            document_type="CREDIT NOTE" if category == "credit_note" else
                          "DEBIT NOTE" if category == "debit_note" else str(document_type).upper(),
            document_number=str(_pick(row, *DOCUMENT_NUMBER_HEADERS) or ""),
            document_date=str(date_value or ""), taxable=values.taxable, igst=values.igst,
            cgst=values.cgst, sgst=values.sgst, cess=values.cess,
            document_value=document_value * sign,
            reverse_charge="Yes" if reverse else ("No" if raw_reverse is not None else "-"),
            itc_availability="Yes",
            period_2a=str(_pick(row, "2a period", "gstr 1 iff gstr 5 period", "gstr1 iff gstr5 period") or ""),
            period_2b=str(period or fallback_period or ""),
            remarks=str(_pick(row, "remarks", "reason", "itc availability reason") or ""),
        ))


def discover_files(folder: str | Path, extract_to: str | Path) -> list[Path]:
    """List every downloaded file, unpacking GSTN's ZIP archives outside the tree.

    Archives are expanded into a scratch directory rather than into the client
    folder so repeated runs never accumulate extra levels inside it.
    """
    folder = Path(folder)
    extract_to = Path(extract_to)
    files: list[Path] = []
    for path in sorted(folder.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() == ".zip":
            target = extract_to / path.relative_to(folder).with_suffix("")
            try:
                with zipfile.ZipFile(path) as archive:
                    archive.extractall(target)
            except zipfile.BadZipFile:
                continue
        elif path.name != "download_status.txt":
            files.append(path)
    files.extend(sorted(path for path in extract_to.rglob("*") if path.is_file()))
    return files


def _period_hint(path: Path) -> str:
    """Recover the return period a downloaded file belongs to.

    Downloads are stored as ``Apr-2025_GSTR2B_<portal name>``; GSTN's own names
    embed the period as ``042025`` or ``2025-04``.
    """
    for candidate in (path.name, path.parent.name):
        month = month_name("", candidate.split("_")[0])
        if month:
            return candidate.split("_")[0]
    match = re.search(r"(?<!\d)(0[1-9]|1[0-2])(20\d{2})(?!\d)", path.name)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return ""


def _report_kind(path: Path) -> str:
    """Decide which return a file belongs to from its folder and file name."""
    haystack = f"{path.parent.name} {path.name}".lower().replace(" ", "")
    for token, kind in (("e-invoice", "einvoice"), ("einvoice", "einvoice"), ("irn", "einvoice"),
                        ("gstr-2b", "2b"), ("gstr2b", "2b"), ("_2b", "2b"),
                        ("gstr-3b", "3b"), ("gstr3b", "3b"),
                        ("gstr-1", "1"), ("gstr1", "1"), ("r1", "1")):
        if token in haystack:
            return kind
    return ""


def parse_download_folder(folder: str | Path) -> tuple[dict[str, MonthData], list[Invoice2B], list[str]]:
    months: dict[str, MonthData] = defaultdict(MonthData)
    invoices: list[Invoice2B] = []
    used: list[str] = []
    with tempfile.TemporaryDirectory(prefix="gstr_extract_") as scratch:
        discovered = discover_files(folder, scratch)
        # Machine-readable downloads are read first so a PDF is only ever used
        # for a month that JSON or Excel did not already cover.
        for path in sorted(discovered, key=lambda item: item.suffix.lower() == ".pdf"):
            suffix = path.suffix.lower()
            kind = _report_kind(path)
            period = _period_hint(path)
            try:
                if suffix == ".json":
                    text = path.read_text(encoding="utf-8-sig", errors="ignore")[:20000].lower()
                    if "sup_details" in text or "itc_elg" in text:
                        parse_gstr3b_json(path, months); used.append(path.name)
                    elif '"b2b"' in text or '"b2cs"' in text or '"b2cl"' in text:
                        parse_gstr1_json(path, months); used.append(path.name)
                elif suffix in {".xlsx", ".xlsm", ".xls"}:
                    if kind == "2b":
                        parse_2b_excel(path, months, invoices, period); used.append(path.name)
                    elif kind == "einvoice":
                        parse_einvoice_excel(path, months, period); used.append(path.name)
                elif suffix == ".pdf":
                    # A filed GSTR-3B has no machine-readable download, so its PDF
                    # is the only source. GSTR-1 PDFs are a fallback used only when
                    # the filed JSON for that month is missing.
                    month = month_name("", period)
                    if kind == "3b" and not (month and months[month].gstr3b):
                        parse_gstr3b_pdf(path, months, period); used.append(path.name)
                    elif kind == "1" and not (month and months[month].gstr1):
                        parse_gstr1_summary_pdf(path, months, period); used.append(path.name)
            except Exception as exc:  # one malformed download must not block the rest
                used.append(f"SKIPPED {path.name}: {exc}")
    return months, invoices, used
