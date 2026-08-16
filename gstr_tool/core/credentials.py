from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import GstCredential


ALIASES = {
    "label": {"client", "client name", "name", "company", "legal name", "trade name"},
    "username": {"user id", "userid", "user", "username", "gst user id", "login id"},
    "password": {"password", "pass", "pwd"},
    "gstin": {"gstin", "gst no", "gst number", "gstin/uin"},
}


def _key(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _find_header(rows: Iterable[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(rows, start=1):
        normalized = [_key(value) for value in row]
        mapping: dict[str, int] = {}
        for field, aliases in ALIASES.items():
            for index, value in enumerate(normalized):
                if value in aliases:
                    mapping[field] = index
                    break
        if {"username", "password"}.issubset(mapping):
            return row_number, mapping
    raise ValueError("Credential workbook needs User ID and Password columns.")


def load_credentials(path: str | Path) -> list[GstCredential]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row, columns = _find_header(rows)
    result: list[GstCredential] = []
    for row in rows[header_row:]:
        username = str(row[columns["username"]] or "").strip()
        password = str(row[columns["password"]] or "")
        if not username or not password:
            continue
        gstin = str(row[columns["gstin"]] or "").strip() if "gstin" in columns else ""
        label = str(row[columns["label"]] or "").strip() if "label" in columns else ""
        result.append(GstCredential(label=label or gstin or username, username=username, password=password, gstin=gstin))
    workbook.close()
    if not result:
        raise ValueError("No credential rows containing both User ID and Password were found.")
    return result
