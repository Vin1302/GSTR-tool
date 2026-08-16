from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from .helpers import MONTHS
from .models import Invoice2B, MonthData, TaxAmounts


def _n(value) -> float:
    return float(value)


def _write_amounts(sheet, row: int, start_col: int, amounts: TaxAmounts, include_taxable: bool = True) -> None:
    values = [amounts.taxable, amounts.igst, amounts.cgst, amounts.sgst, amounts.cess] if include_taxable else [amounts.igst, amounts.cgst, amounts.sgst, amounts.cess]
    for offset, value in enumerate(values):
        sheet.cell(row, start_col + offset).value = _n(value)


def _get(data: dict[str, TaxAmounts], key: str) -> TaxAmounts:
    return data.get(key, TaxAmounts())


def _write_3b(sheet, months: dict[str, MonthData]) -> None:
    for row, month in enumerate(MONTHS, start=7):
        data = months.get(month, MonthData()).gstr3b
        outward = _get(data, "outward_nrc")
        non_taxable = _get(data, "non_taxable")
        rcm = _get(data, "outward_rcm")
        sheet.cell(row, 2).value = _n(outward.taxable)
        sheet.cell(row, 3).value = _n(non_taxable.taxable)
        _write_amounts(sheet, row, 6, outward, include_taxable=False)
        sheet.cell(row, 12).value = _n(rcm.taxable)
        _write_amounts(sheet, row, 15, rcm, include_taxable=False)
        _write_amounts(sheet, row + 17, 2, _get(data, "itc_other"), include_taxable=False)
        _write_amounts(sheet, row + 17, 6, _get(data, "itc_reversed"), include_taxable=False)
        _write_amounts(sheet, row + 53, 2, _get(data, "itc_rcm"), include_taxable=False)
        _write_amounts(sheet, row + 53, 18, _get(data, "late_fee"), include_taxable=False)


def _write_einvoice(sheet, months: dict[str, MonthData]) -> None:
    for row, month in enumerate(MONTHS, start=6):
        data = months.get(month, MonthData()).einvoice
        for start_col, category in ((2, "b2b"), (6, "debit_note"), (10, "credit_note")):
            amounts = _get(data, category)
            for offset, value in enumerate((amounts.taxable, amounts.igst, amounts.cgst, amounts.sgst)):
                sheet.cell(row, start_col + offset).value = _n(value)


def _write_gstr1(sheet, months: dict[str, MonthData]) -> None:
    mapping = {
        "b2b": 2, "b2b_rcm": 7, "b2c": 12, "b2b_amended": 17,
        "b2c_amended": 22, "cdn": 27, "hsn": 43,
    }
    for row, month in enumerate(MONTHS, start=6):
        data = months.get(month, MonthData()).gstr1
        for category, start_col in mapping.items():
            _write_amounts(sheet, row, start_col, _get(data, category), include_taxable=True)


def _write_2b(sheet, months: dict[str, MonthData]) -> None:
    for row, month in enumerate(MONTHS, start=6):
        data = months.get(month, MonthData()).gstr2b
        _write_amounts(sheet, row, 2, _get(data, "other"), include_taxable=True)
        _write_amounts(sheet, row, 7, _get(data, "rcm"), include_taxable=True)


def _write_invoice_2b(sheet, invoices: list[Invoice2B]) -> None:
    start_row = 5
    if sheet.max_row >= start_row:
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=15):
            for cell in row:
                cell.value = None
    style_source_row = start_row
    for index, invoice in enumerate(invoices, start=start_row):
        if index > style_source_row:
            for column in range(1, 16):
                source = sheet.cell(style_source_row, column)
                target = sheet.cell(index, column)
                target._style = copy(source._style)
                target.number_format = source.number_format
        values = [
            invoice.party, invoice.gstin, invoice.document_type, invoice.document_number,
            invoice.document_date, _n(invoice.taxable), _n(invoice.igst),
            _n(invoice.cgst + invoice.sgst), _n(invoice.cess), _n(invoice.document_value),
            invoice.reverse_charge, invoice.itc_availability, invoice.period_2a,
            invoice.period_2b, invoice.remarks,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(index, column).value = value


def generate_workbook(template_path: str | Path, output_path: str | Path,
                      months: dict[str, MonthData], invoices: list[Invoice2B]) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    workbook = load_workbook(output_path, data_only=False)
    required = {"As per 3B", "As per E-Invoice", "As per GSTR 1", "GSTR 2b", "Invoice Wise 2B(incl CDNR)"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Template is missing required sheets: {', '.join(sorted(missing))}")
    _write_3b(workbook["As per 3B"], months)
    _write_einvoice(workbook["As per E-Invoice"], months)
    _write_gstr1(workbook["As per GSTR 1"], months)
    _write_2b(workbook["GSTR 2b"], months)
    _write_invoice_2b(workbook["Invoice Wise 2B(incl CDNR)"], invoices)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(output_path)
    workbook.close()
    return output_path
